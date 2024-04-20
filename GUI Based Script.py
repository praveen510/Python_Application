import os
import csv
import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import time

def replace_text_csv(filepath, old_text, new_text, progress_bar, status_label, timer_label):
    try:
        start_time = time.time()
        total_size = os.path.getsize(filepath)
        read_size = 0
        
        with open(filepath, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            rows = list(reader)
        
        total_cells = sum(len(row) for row in rows)
        replaced_cells = 0
        
        with open(filepath, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for row in rows:
                new_row = [cell.replace(old_text, new_text) if old_text in cell else cell for cell in row]
                writer.writerow(new_row)
                replaced_cells += len([cell for cell in new_row if old_text in cell])
                read_size += csvfile.tell() - read_size
                progress = (read_size / total_size) * 100
                progress_bar['value'] = progress
                window.update_idletasks()
                elapsed_time = time.time() - start_time
                remaining_time = (total_size - read_size) / read_size * elapsed_time if read_size > 0 else 0
                timer_label.config(text="Time left: {:.2f} seconds".format(remaining_time))
        
        status_label.config(text="Successfully replaced '{}' with '{}' in the CSV file.".format(old_text, new_text))
    except Exception as e:
        status_label.config(text="An error occurred: {}".format(str(e)))

# def replace_text_excel(filepath, old_text, new_text, progress_bar, status_label, timer_label):
#     try:
#         start_time = time.time()
#         replaced_cells = 0

#         # Open the workbook
#         wb = openpyxl.load_workbook(filepath)
#         total_cells = sum(sheet.max_row * sheet.max_column for sheet in wb.worksheets)

#         # Iterate over all sheets in the workbook
#         for sheet in wb.worksheets:
#             for row in sheet.iter_rows(values_only=True):
#                 for cell_value in row:
#                     if cell_value and old_text in str(cell_value):
#                         # Replace old text with new text
#                         new_value = str(cell_value).replace(old_text, new_text)
#                         # Update the cell with new value
#                         sheet.cell(row=row[0].row, column=row.index(cell_value) + 1, value=new_value)
#                         replaced_cells += 1
#                         progress = (replaced_cells / total_cells) * 100
#                         progress_bar['value'] = progress
#                         window.update_idletasks()
#                         elapsed_time = time.time() - start_time
#                         remaining_time = (total_cells - replaced_cells) / replaced_cells * elapsed_time if replaced_cells > 0 else 0
#                         timer_label.config(text="Time left: {:.2f} seconds".format(remaining_time))

#         # Save the workbook
#         wb.save(filepath)

#         # Update status label
#         status_label.config(text="Successfully replaced '{}' with '{}' in the Excel file.".format(old_text, new_text))
#     except Exception as e:
#         # Update status label with error message
#         status_label.config(text="An error occurred: {}".format(str(e)))

def select_file():
    file_path = filedialog.askopenfilename()
    return file_path

def replace_csv():
    file_path = select_file()
    if file_path:
        old_text = entry_old.get()
        new_text = entry_new.get()
        progress_bar = ttk.Progressbar(window, orient=tk.HORIZONTAL, length=200, mode='determinate')
        progress_bar.pack()
        status_label = ttk.Label(window, text="", anchor=tk.W)
        status_label.pack(fill=tk.X)
        timer_label = ttk.Label(window, text="")
        timer_label.pack(fill=tk.X)
        replace_text_csv(file_path, old_text, new_text, progress_bar, status_label, timer_label)
        progress_bar.destroy()
        status_label.destroy()
        timer_label.destroy()

# def replace_excel():
#     file_path = select_file()
#     if file_path:
#         old_text = entry_old.get()
#         new_text = entry_new.get()
#         progress_bar = ttk.Progressbar(window, orient=tk.HORIZONTAL, length=200, mode='determinate')
#         progress_bar.pack()
#         status_label = ttk.Label(window, text="", anchor=tk.W)
#         status_label.pack(fill=tk.X)
#         timer_label = ttk.Label(window, text="")
#         timer_label.pack(fill=tk.X)
#         replace_text_excel(file_path, old_text, new_text, progress_bar, status_label, timer_label)
#         progress_bar.destroy()
#         status_label.destroy()
#         timer_label.destroy()

# Create the main window
window = tk.Tk()
window.title("Replace Text in CSV")
window.geometry("600x550")

# Create input fields
label_old = ttk.Label(window, text="Old Text:", font=("Arial", 12))  # Increased font size
label_old.pack()
entry_old = ttk.Entry(window, font=("Arial", 12))  # Increased font size
entry_old.pack()

label_new = ttk.Label(window, text="New Text:", font=("Arial", 12))  # Increased font size
label_new.pack()
entry_new = ttk.Entry(window, font=("Arial", 12))  # Increased font size
entry_new.pack()

# Create a style for buttons
style = ttk.Style()
style.configure('My.TButton', font=('Arial', 12))

# Create buttons with increased text size
button_csv = ttk.Button(window, text="Replace in CSV", width=20, command=replace_csv, style='My.TButton')  # Increased font size
button_csv.pack()

# button_excel = ttk.Button(window, text="Replace in Excel", width=20, command=replace_excel, style='My.TButton')  # Increased font size
# button_excel.pack()

# Run the main loop
window.mainloop()
